#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
同步 household_ratio.xlsm -> household_db.json
使用方式:
  python sync_household.py
"""
import json, os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("請安裝 openpyxl: pip install openpyxl")
    raise

BASE = Path(__file__).parent
EXCEL = BASE / "household_ratio.xlsm"
JSON_PATH = BASE / "household_db.json"

def load_excel():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb.active
    # 取得標題定位
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    house_idx = next((i for i,h in enumerate(header) if h and "戶號" in str(h)), 0)
    share_idx = next((i for i,h in enumerate(header) if h and ("持分比" in str(h) or "持分" in str(h))), 2)
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        house = row[house_idx]
        share = row[share_idx]
        if house and share is not None:
            try:
                data[str(house).strip()] = f"{float(share):.3f}"
            except Exception:
                continue
    return data

def main():
    new_data = load_excel()
    # 讀舊檔
    old = {}
    if JSON_PATH.exists():
        old = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    merged = {**old, **new_data}
    JSON_PATH.write_text(json.dumps(merged, ensure_ascii=False), encoding="utf-8")
    print(f"同步完成: {len(new_data)} 條 Excel -> {len(merged)} 條 JSON 已寫入 {JSON_PATH.name}")

if __name__ == "__main__":
    main()
